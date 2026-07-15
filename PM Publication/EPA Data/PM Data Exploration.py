# ============================================================
# EPA AIR QUALITY EXCEL EXPLORATION SCRIPT
# Robust version
#
# Purpose:
#   Open every EPA air quality Excel workbook.
#   Explore every sheet.
#   Detect possible two-level headers:
#       row 1 = variable names
#       row 2 = units
#   Export a structured exploration report.
#
# Main output:
#   PM_Data_EPA/exploration_outputs/EPA_air_quality_structure_exploration_report.xlsx
# ============================================================


# ============================================================
# 0. IMPORTS
# ============================================================

import os
import re
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.simplefilter("ignore", category=UserWarning)


# ============================================================
# 1. USER SETTINGS
# ============================================================

# Main folder where the EPA files should be
DATA_DIR = Path(r"C:/Users/Richa/Documents/Python_Projects/PM_Data_EPA")

# If Python somehow cannot find the files in DATA_DIR, it will also check
# the current working directory and nearby folders.
ALSO_SEARCH_CURRENT_WORKING_DIRECTORY = True
SEARCH_RECURSIVELY = False

OUTPUT_DIR = DATA_DIR / "exploration_outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

REPORT_XLSX = OUTPUT_DIR / "EPA_air_quality_structure_exploration_report.xlsx"

MAX_PREVIEW_ROWS = 25
MAX_HEADER_SCAN_ROWS = 20

VALID_EXCEL_SUFFIXES = [".xlsm", ".xlsx", ".xls", ".xlms"]


# ============================================================
# 2. GENERAL HELPER FUNCTIONS
# ============================================================

def clean_text(x):
    """
    Clean a cell value into a readable string.
    """
    if pd.isna(x):
        return ""
    x = str(x).strip()
    x = re.sub(r"\s+", " ", x)
    return x


def safe_sheet_name(name, max_len=31):
    """
    Excel sheet names have a 31-character limit and cannot contain certain characters.
    """
    name = str(name)
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    return name[:max_len]


def print_folder_diagnostic(folder):
    """
    Print exactly what Python can see in a folder.
    """
    folder = Path(folder)

    print("\n============================================================")
    print("FOLDER DIAGNOSTIC")
    print("============================================================")
    print("Folder:", folder)
    print("Folder exists:", folder.exists())

    if not folder.exists():
        print("Folder does not exist.")
        return

    print("\nItems Python can see:")
    for f in sorted(folder.iterdir()):
        print(f"  name={repr(f.name)} | suffix={repr(f.suffix)} | is_file={f.is_file()}")


def find_excel_files(data_dir):
    """
    Find Excel files robustly.
    First checks DATA_DIR.
    If none found, optionally checks current working directory.
    """
    data_dir = Path(data_dir)

    print_folder_diagnostic(data_dir)

    search_folders = [data_dir]

    if ALSO_SEARCH_CURRENT_WORKING_DIRECTORY:
        cwd = Path.cwd()
        if cwd not in search_folders:
            search_folders.append(cwd)

    files = []

    for folder in search_folders:
        if not folder.exists():
            continue

        print("\nSearching folder:")
        print(folder)

        if SEARCH_RECURSIVELY:
            candidates = list(folder.rglob("*"))
        else:
            candidates = list(folder.iterdir())

        for f in candidates:
            if not f.is_file():
                continue

            suffix = f.suffix.lower().strip()

            # Accept normal Excel suffixes
            if suffix in VALID_EXCEL_SUFFIXES:
                files.append(f)

            # Also accept files that contain air_quality_Ireland anywhere in name
            # even if suffix is strange
            elif "air_quality_ireland" in f.name.lower():
                files.append(f)

    # Remove duplicates while preserving order
    unique_files = []
    seen = set()

    for f in files:
        resolved = str(f.resolve()).lower()
        if resolved not in seen:
            unique_files.append(f)
            seen.add(resolved)

    files = sorted(unique_files)

    print("\n============================================================")
    print("EXCEL FILE SEARCH RESULT")
    print("============================================================")

    if files:
        print("Files found:")
        for f in files:
            print(f"  {f}")
    else:
        print("No files found.")

    if not files:
        raise FileNotFoundError(
            "\nNo Excel files were found.\n\n"
            f"Primary folder checked: {data_dir}\n"
            f"Current working directory: {Path.cwd()}\n"
            f"Accepted suffixes: {VALID_EXCEL_SUFFIXES}\n\n"
            "Run the folder diagnostic printed above and check whether Python is seeing the same folder as Windows Explorer."
        )

    return files


def workbook_metadata(path):
    """
    Extract workbook-level metadata using openpyxl.
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=False, keep_vba=True)
    except Exception as e:
        return [{
            "file": path.name,
            "full_path": str(path),
            "sheet": None,
            "max_row_openpyxl": None,
            "max_column_openpyxl": None,
            "workbook_read_status": "failed",
            "error": str(e)
        }]

    rows = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows.append({
            "file": path.name,
            "full_path": str(path),
            "sheet": sheet,
            "max_row_openpyxl": ws.max_row,
            "max_column_openpyxl": ws.max_column,
            "workbook_read_status": "success",
            "error": ""
        })

    wb.close()
    return rows


def read_raw_sheet(path, sheet_name, nrows=None):
    """
    Read sheet without assuming headers.
    Everything is read as object initially to preserve structure.
    """
    return pd.read_excel(
        path,
        sheet_name=sheet_name,
        header=None,
        nrows=nrows,
        engine="openpyxl"
    )


def row_nonempty_count(row):
    """
    Count non-empty cells in a row.
    """
    return sum(clean_text(x) != "" for x in row)


def row_text_fraction(row):
    """
    Estimate how much of a row is text-like.
    """
    vals = [clean_text(x) for x in row]
    vals = [v for v in vals if v != ""]

    if len(vals) == 0:
        return 0

    text_like = 0

    for v in vals:
        numeric = pd.to_numeric(pd.Series([v]), errors="coerce").notna().iloc[0]
        datetime_like = pd.to_datetime(pd.Series([v]), errors="coerce", dayfirst=True).notna().iloc[0]

        if not numeric and not datetime_like:
            text_like += 1

    return text_like / len(vals)


def detect_header_rows(raw_df, max_scan_rows=20):
    """
    Detect likely header row and possible unit row.

    EPA files often have:
        row n     = variable names
        row n + 1 = units
        row n + 2 = data start

    This function is intentionally conservative.
    """
    scan = raw_df.head(max_scan_rows).copy()

    row_info = []

    for idx in scan.index:
        row = scan.loc[idx]
        nonempty = row_nonempty_count(row)
        text_frac = row_text_fraction(row)

        row_info.append({
            "row_index": idx,
            "nonempty": nonempty,
            "text_fraction": text_frac
        })

    row_info_df = pd.DataFrame(row_info)

    # Candidate header rows:
    # - at least 2 non-empty cells
    # - some text-like values
    candidates = row_info_df[
        (row_info_df["nonempty"] >= 2) &
        (row_info_df["text_fraction"] >= 0.3)
    ]

    if candidates.empty:
        header_row = 0
    else:
        # Pick the first plausible header row
        header_row = int(candidates.iloc[0]["row_index"])

    units_row = None
    header_type = "single_header"

    # Check row below header for units
    if header_row + 1 < len(raw_df):
        possible_units = raw_df.iloc[header_row + 1].astype(str).map(clean_text)
        joined = " ".join(possible_units.tolist()).lower()

        unit_patterns = [
            r"µg", r"ug", r"μg", r"mg", r"g/m", r"m3", r"m\^3",
            r"°c", r"deg", r"%", r"hpa", r"mb",
            r"m/s", r"km/h", r"mm", r"w/m",
            r"ppm", r"ppb", r"µmol", r"umol",
            r"unit", r"units", r"count", r"number"
        ]

        unit_hits = sum(1 for pat in unit_patterns if re.search(pat, joined))

        possible_units_nonempty = sum(
            v != "" and v.lower() not in ["nan", "none"]
            for v in possible_units
        )

        # Detect as unit row if unit-like text appears
        if unit_hits >= 1 and possible_units_nonempty >= 1:
            units_row = header_row + 1
            header_type = "two_level_header_variable_plus_units"

    data_start_row = header_row + 1 if units_row is None else units_row + 1

    return {
        "header_row": int(header_row),
        "units_row": None if units_row is None else int(units_row),
        "data_start_row": int(data_start_row),
        "header_type": header_type
    }


def make_columns_from_headers(raw_df, header_row, units_row=None):
    """
    Build clean column names from one or two header rows.

    If units row exists:
        variable [unit]
    """
    variables = raw_df.iloc[header_row].tolist()
    variables = [clean_text(x) for x in variables]

    if units_row is not None:
        units = raw_df.iloc[units_row].tolist()
        units = [clean_text(x) for x in units]
    else:
        units = [""] * len(variables)

    columns = []
    seen = {}

    for i, (var, unit) in enumerate(zip(variables, units)):
        if var == "" or var.lower() in ["nan", "none"]:
            var = f"unnamed_col_{i + 1}"

        if unit != "" and unit.lower() not in ["nan", "none", "unit", "units"]:
            col = f"{var} [{unit}]"
        else:
            col = var

        col = clean_text(col)

        if col in seen:
            seen[col] += 1
            col = f"{col}_{seen[col]}"
        else:
            seen[col] = 1

        columns.append(col)

    return columns


def read_structured_sheet(path, sheet_name, header_info):
    """
    Read the sheet using detected header structure.
    """
    raw = read_raw_sheet(path, sheet_name)

    columns = make_columns_from_headers(
        raw_df=raw,
        header_row=header_info["header_row"],
        units_row=header_info["units_row"]
    )

    df = raw.iloc[header_info["data_start_row"]:].copy()
    df.columns = columns

    # Drop fully empty rows and fully empty columns
    df = df.dropna(axis=0, how="all")
    df = df.dropna(axis=1, how="all")

    df = df.reset_index(drop=True)

    return df


def infer_column_roles(df):
    """
    Identify likely date, time, PM, pollutant, station, and meteorological columns.
    """
    rows = []

    for col in df.columns:
        col_lower = str(col).lower()

        role = "unknown"

        if any(x in col_lower for x in ["datetime", "date", "time", "hour", "day"]):
            role = "time_date"
        elif any(x in col_lower for x in ["pm2.5", "pm 2.5", "pm25", "pm_2_5", "pm_25"]):
            role = "pm2_5"
        elif any(x in col_lower for x in ["pm10", "pm 10", "pm_10"]):
            role = "pm10"
        elif any(x in col_lower for x in ["no2", "nitrogen dioxide"]):
            role = "no2"
        elif any(x in col_lower for x in ["nox", "nitrogen oxides"]):
            role = "nox"
        elif any(x in col_lower for x in [" o3", "o3 ", "ozone"]):
            role = "ozone"
        elif any(x in col_lower for x in ["so2", "sulphur dioxide", "sulfur dioxide"]):
            role = "so2"
        elif any(x in col_lower for x in [" co ", "carbon monoxide"]):
            role = "co"
        elif any(x in col_lower for x in ["temp", "temperature"]):
            role = "temperature"
        elif any(x in col_lower for x in ["rh", "humidity"]):
            role = "humidity"
        elif any(x in col_lower for x in ["wind"]):
            role = "wind"
        elif any(x in col_lower for x in ["rain", "precip"]):
            role = "precipitation"
        elif any(x in col_lower for x in ["pressure", "hpa", "mb"]):
            role = "pressure"
        elif any(x in col_lower for x in ["station", "site", "location", "county", "area", "zone"]):
            role = "site_metadata"

        rows.append({
            "column": col,
            "inferred_role": role
        })

    return pd.DataFrame(rows)


def summarize_dataframe(df, file_name, sheet_name):
    """
    Produce column-level summary for a structured dataframe.
    """
    summary_rows = []

    for col in df.columns:
        s = df[col]

        n_total = len(s)
        n_missing = int(s.isna().sum())
        n_non_missing = int(s.notna().sum())

        numeric_s = pd.to_numeric(s, errors="coerce")
        n_numeric = int(numeric_s.notna().sum())
        numeric_fraction = n_numeric / max(n_non_missing, 1)

        datetime_s = pd.to_datetime(s, errors="coerce", dayfirst=True)
        n_datetime = int(datetime_s.notna().sum())
        datetime_fraction = n_datetime / max(n_non_missing, 1)

        likely_type = "text"

        if n_non_missing == 0:
            likely_type = "empty"
        elif numeric_fraction >= 0.8:
            likely_type = "numeric"
        elif datetime_fraction >= 0.8:
            likely_type = "datetime"

        examples = (
            s.dropna()
            .astype(str)
            .head(8)
            .tolist()
        )

        row = {
            "file": file_name,
            "sheet": sheet_name,
            "column": col,
            "rows_total": n_total,
            "non_missing": n_non_missing,
            "missing": n_missing,
            "missing_percent": round(100 * n_missing / max(n_total, 1), 2),
            "raw_dtype": str(s.dtype),
            "likely_type": likely_type,
            "n_unique": int(s.nunique(dropna=True)),
            "example_values": " | ".join(examples)
        }

        if likely_type == "numeric":
            row.update({
                "numeric_min": numeric_s.min(),
                "numeric_q25": numeric_s.quantile(0.25),
                "numeric_median": numeric_s.median(),
                "numeric_mean": numeric_s.mean(),
                "numeric_q75": numeric_s.quantile(0.75),
                "numeric_max": numeric_s.max(),
            })
        else:
            row.update({
                "numeric_min": np.nan,
                "numeric_q25": np.nan,
                "numeric_median": np.nan,
                "numeric_mean": np.nan,
                "numeric_q75": np.nan,
                "numeric_max": np.nan,
            })

        if likely_type == "datetime":
            row.update({
                "datetime_min": datetime_s.min(),
                "datetime_max": datetime_s.max(),
            })
        else:
            row.update({
                "datetime_min": pd.NaT,
                "datetime_max": pd.NaT,
            })

        summary_rows.append(row)

    return pd.DataFrame(summary_rows)


def inspect_top_rows(raw_df, file_name, sheet_name):
    """
    Store the first rows exactly as read, useful for seeing messy headings.
    """
    top = raw_df.head(MAX_PREVIEW_ROWS).copy()
    top.insert(0, "row_number_excel_like", np.arange(1, len(top) + 1))
    top.insert(0, "sheet", sheet_name)
    top.insert(0, "file", file_name)
    return top


def save_csv_preview(df, file_stem, sheet_name):
    """
    Save a preview of each parsed sheet.
    """
    safe_name = re.sub(r"[^A-Za-z0-9_]+", "_", f"{file_stem}_{sheet_name}")
    out_path = OUTPUT_DIR / f"preview_{safe_name[:120]}.csv"
    df.head(150).to_csv(out_path, index=False)
    return out_path


def make_sheet_header_debug(raw_df, file_name, sheet_name, header_info):
    """
    Save information about the first rows and chosen header rows.
    """
    rows = []

    n = min(MAX_HEADER_SCAN_ROWS, len(raw_df))

    for i in range(n):
        values = [clean_text(x) for x in raw_df.iloc[i].tolist()]
        values_short = " | ".join(values[:20])

        rows.append({
            "file": file_name,
            "sheet": sheet_name,
            "excel_row_number": i + 1,
            "nonempty_count": row_nonempty_count(raw_df.iloc[i]),
            "text_fraction": row_text_fraction(raw_df.iloc[i]),
            "selected_as_header": i == header_info["header_row"],
            "selected_as_units": False if header_info["units_row"] is None else i == header_info["units_row"],
            "selected_as_data_start": i == header_info["data_start_row"],
            "row_values_first_20_columns": values_short
        })

    return pd.DataFrame(rows)


# ============================================================
# 3. MAIN EXPLORATION
# ============================================================

excel_files = find_excel_files(DATA_DIR)

workbook_rows = []
sheet_detection_rows = []
column_summary_list = []
column_role_list = []
top_rows_list = []
parsed_preview_list = []
header_debug_list = []

print("\n============================================================")
print("STARTING WORKBOOK EXPLORATION")
print("============================================================")

for path in excel_files:

    print("\n============================================================")
    print(f"WORKBOOK: {path.name}")
    print(f"FULL PATH: {path}")
    print("============================================================")

    # Get workbook metadata
    meta_rows = workbook_metadata(path)
    workbook_rows.extend(meta_rows)

    # Try opening workbook
    try:
        wb = load_workbook(path, read_only=True, data_only=False, keep_vba=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        print(f"Could not open workbook: {path.name}")
        print(e)

        sheet_detection_rows.append({
            "file": path.name,
            "full_path": str(path),
            "sheet": None,
            "status": "failed_to_open_workbook",
            "error": str(e)
        })
        continue

    print(f"Number of sheets: {len(sheet_names)}")
    print("Sheets:")
    for sheet in sheet_names:
        print(f"  - {sheet}")

    for sheet_name in sheet_names:

        print("\n------------------------------------------------------------")
        print(f"Reading sheet: {sheet_name}")
        print("------------------------------------------------------------")

        try:
            raw = read_raw_sheet(path, sheet_name)
        except Exception as e:
            print(f"Could not read raw sheet: {sheet_name}")
            print(e)

            sheet_detection_rows.append({
                "file": path.name,
                "full_path": str(path),
                "sheet": sheet_name,
                "status": "failed_to_read_raw_sheet",
                "error": str(e)
            })
            continue

        # Save top rows
        top_rows_list.append(
            inspect_top_rows(
                raw_df=raw,
                file_name=path.name,
                sheet_name=sheet_name
            )
        )

        # Detect header structure
        header_info = detect_header_rows(raw, MAX_HEADER_SCAN_ROWS)

        header_debug_list.append(
            make_sheet_header_debug(
                raw_df=raw,
                file_name=path.name,
                sheet_name=sheet_name,
                header_info=header_info
            )
        )

        # Parse structured sheet
        try:
            df = read_structured_sheet(path, sheet_name, header_info)
        except Exception as e:
            print(f"Could not parse structured sheet: {sheet_name}")
            print(e)

            sheet_detection_rows.append({
                "file": path.name,
                "full_path": str(path),
                "sheet": sheet_name,
                "status": "failed_to_parse_structured_sheet",
                "error": str(e),
                "header_type": header_info["header_type"],
                "header_row_excel_like": header_info["header_row"] + 1,
                "units_row_excel_like": None if header_info["units_row"] is None else header_info["units_row"] + 1,
                "data_start_row_excel_like": header_info["data_start_row"] + 1,
            })
            continue

        preview_path = save_csv_preview(df, path.stem, sheet_name)

        sheet_detection_rows.append({
            "file": path.name,
            "full_path": str(path),
            "sheet": sheet_name,
            "status": "parsed",
            "error": "",
            "header_type": header_info["header_type"],
            "header_row_excel_like": header_info["header_row"] + 1,
            "units_row_excel_like": None if header_info["units_row"] is None else header_info["units_row"] + 1,
            "data_start_row_excel_like": header_info["data_start_row"] + 1,
            "n_rows_raw": raw.shape[0],
            "n_cols_raw": raw.shape[1],
            "n_rows_parsed": df.shape[0],
            "n_cols_parsed": df.shape[1],
            "preview_csv": str(preview_path)
        })

        print(f"Detected header type: {header_info['header_type']}")
        print(f"Header row: {header_info['header_row'] + 1}")
        print(f"Units row: {None if header_info['units_row'] is None else header_info['units_row'] + 1}")
        print(f"Data starts row: {header_info['data_start_row'] + 1}")
        print(f"Parsed shape: {df.shape}")
        print("First 15 parsed columns:")
        print(list(df.columns[:15]))

        # Column-level summary
        col_sum = summarize_dataframe(df, path.name, sheet_name)
        column_summary_list.append(col_sum)

        # Column roles
        roles = infer_column_roles(df)
        roles.insert(0, "sheet", sheet_name)
        roles.insert(0, "file", path.name)
        column_role_list.append(roles)

        # Parsed preview
        preview = df.head(MAX_PREVIEW_ROWS).copy()
        preview.insert(0, "sheet", sheet_name)
        preview.insert(0, "file", path.name)
        parsed_preview_list.append(preview)


# ============================================================
# 4. COMBINE RESULTS
# ============================================================

workbook_summary = pd.DataFrame(workbook_rows)
sheet_detection = pd.DataFrame(sheet_detection_rows)

if column_summary_list:
    column_summary = pd.concat(column_summary_list, ignore_index=True)
else:
    column_summary = pd.DataFrame()

if column_role_list:
    column_roles = pd.concat(column_role_list, ignore_index=True)
else:
    column_roles = pd.DataFrame()

if top_rows_list:
    top_raw_rows = pd.concat(top_rows_list, ignore_index=True, sort=False)
else:
    top_raw_rows = pd.DataFrame()

if parsed_preview_list:
    parsed_previews = pd.concat(parsed_preview_list, ignore_index=True, sort=False)
else:
    parsed_previews = pd.DataFrame()

if header_debug_list:
    header_debug = pd.concat(header_debug_list, ignore_index=True, sort=False)
else:
    header_debug = pd.DataFrame()


# ============================================================
# 5. CROSS-WORKBOOK COMPARISONS
# ============================================================

if not column_summary.empty:
    column_presence = (
        column_summary
        .groupby(["sheet", "column"], dropna=False)
        .agg(
            n_files_present=("file", "nunique"),
            files_present=("file", lambda x: " | ".join(sorted(set(x)))),
            likely_types=("likely_type", lambda x: " | ".join(sorted(set(map(str, x)))))
        )
        .reset_index()
        .sort_values(["sheet", "column"])
    )
else:
    column_presence = pd.DataFrame()

if not sheet_detection.empty and "status" in sheet_detection.columns:
    parsed_only = sheet_detection[sheet_detection["status"] == "parsed"].copy()

    if not parsed_only.empty:
        sheet_presence = (
            parsed_only
            .groupby("sheet", dropna=False)
            .agg(
                n_files_present=("file", "nunique"),
                files_present=("file", lambda x: " | ".join(sorted(set(x)))),
                parsed_rows_min=("n_rows_parsed", "min"),
                parsed_rows_max=("n_rows_parsed", "max"),
                parsed_cols_min=("n_cols_parsed", "min"),
                parsed_cols_max=("n_cols_parsed", "max"),
            )
            .reset_index()
            .sort_values("sheet")
        )
    else:
        sheet_presence = pd.DataFrame()
else:
    sheet_presence = pd.DataFrame()


# ============================================================
# 6. IDENTIFY LIKELY PM SHEETS AND COLUMNS
# ============================================================

if not column_roles.empty:
    pm_related_columns = column_roles[
        column_roles["inferred_role"].isin(["pm2_5", "pm10"])
    ].copy()
else:
    pm_related_columns = pd.DataFrame()

if not column_summary.empty:
    likely_date_columns = column_summary[
        column_summary["column"].astype(str).str.lower().str.contains(
            "date|time|datetime|hour|day", regex=True, na=False
        )
    ].copy()
else:
    likely_date_columns = pd.DataFrame()


# ============================================================
# 7. WRITE REPORT
# ============================================================

print("\n============================================================")
print("WRITING REPORT")
print("============================================================")
print(REPORT_XLSX)

with pd.ExcelWriter(REPORT_XLSX, engine="openpyxl") as writer:

    workbook_summary.to_excel(writer, sheet_name="workbook_summary", index=False)
    sheet_detection.to_excel(writer, sheet_name="sheet_detection", index=False)

    if not sheet_presence.empty:
        sheet_presence.to_excel(writer, sheet_name="sheet_presence", index=False)

    if not column_presence.empty:
        column_presence.to_excel(writer, sheet_name="column_presence", index=False)

    if not column_summary.empty:
        column_summary.to_excel(writer, sheet_name="column_summary", index=False)

    if not column_roles.empty:
        column_roles.to_excel(writer, sheet_name="column_roles", index=False)

    if not pm_related_columns.empty:
        pm_related_columns.to_excel(writer, sheet_name="pm_related_columns", index=False)

    if not likely_date_columns.empty:
        likely_date_columns.to_excel(writer, sheet_name="likely_date_columns", index=False)

    if not header_debug.empty:
        header_debug.to_excel(writer, sheet_name="header_debug", index=False)

    if not top_raw_rows.empty:
        top_raw_rows.to_excel(writer, sheet_name="top_raw_rows", index=False)

    if not parsed_previews.empty:
        parsed_previews.to_excel(writer, sheet_name="parsed_previews", index=False)


# ============================================================
# 8. FINAL CONSOLE SUMMARY
# ============================================================

print("\n============================================================")
print("EXPLORATION COMPLETE")
print("============================================================")

print(f"\nExcel files explored: {len(excel_files)}")
for f in excel_files:
    print(f"  - {f.name}")

print(f"\nReport written to:")
print(REPORT_XLSX)

print(f"\nCSV previews written to:")
print(OUTPUT_DIR)

print("\nOpen this file first:")
print(REPORT_XLSX)

print("\nImportant report tabs:")
print("1. sheet_detection")
print("   Shows each workbook, each sheet, detected header row, units row, and parsed shape.")
print("2. header_debug")
print("   Shows the first rows of each sheet and which row was chosen as header/units/data start.")
print("3. top_raw_rows")
print("   Shows the raw top rows exactly as Python read them.")
print("4. column_summary")
print("   Shows missingness, likely type, numeric ranges, and example values for every column.")
print("5. column_roles")
print("   Guesses whether columns are PM2.5, PM10, date/time, met variables, site metadata, etc.")
print("6. column_presence")
print("   Shows whether columns are consistent across years.")
print("7. pm_related_columns")
print("   Quickly isolates likely PM2.5 and PM10 columns.")
