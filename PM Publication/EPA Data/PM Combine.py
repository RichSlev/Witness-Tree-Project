# ============================================================
# COMBINE EPA AIR QUALITY PM DATA INTO ONE DAILY FILE
#
# Input folder:
#   C:/Users/Richa/Documents/Python_Projects/PM_Data_EPA
#
# Expected files:
#   2020_air_quality_Ireland.xlsm
#   2021_air_quality_Ireland.xlsm
#   2022_air_quality_Ireland.xlsm
#   2023_air_quality_Ireland.xlsm
#   2024_air_quality_Ireland.xlsm
#
# Output:
#   PM_Data_EPA/combined_outputs/EPA_PM_daily_long_2020_2024.csv
#   PM_Data_EPA/combined_outputs/EPA_PM_daily_long_2020_2024.xlsx
#   PM_Data_EPA/combined_outputs/EPA_PM_daily_station_summary_2020_2024.csv
#   PM_Data_EPA/combined_outputs/EPA_PM_daily_file_sheet_log.csv
# ============================================================


# ============================================================
# 0. IMPORTS
# ============================================================

import re
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.simplefilter("ignore", category=UserWarning)


# ============================================================
# 1. USER SETTINGS
# ============================================================

DATA_DIR = Path(r"C:/Users/Richa/Documents/Python_Projects/PM_Data_EPA")

OUTPUT_DIR = DATA_DIR / "combined_outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

OUT_LONG_CSV = OUTPUT_DIR / "EPA_PM_daily_long_2020_2024.csv"
OUT_LONG_XLSX = OUTPUT_DIR / "EPA_PM_daily_long_2020_2024.xlsx"
OUT_SUMMARY_CSV = OUTPUT_DIR / "EPA_PM_daily_station_summary_2020_2024.csv"
OUT_LOG_CSV = OUTPUT_DIR / "EPA_PM_daily_file_sheet_log.csv"

YEARS_TO_PROCESS = [2020, 2021, 2022, 2023, 2024]

VALID_EXCEL_SUFFIXES = [".xlsm", ".xlsx", ".xls", ".xlms"]

# Set this to True if you want very verbose console output
VERBOSE = True


# ============================================================
# 2. BASIC HELPERS
# ============================================================

def clean_text(x):
    """
    Clean a cell value into a readable string.
    """
    if pd.isna(x):
        return ""
    x = str(x).strip()
    x = re.sub(r"\s+", " ", x)
    return x


def normalise_name(x):
    """
    Lowercase and simplify text for matching.
    """
    x = clean_text(x).lower()
    x = x.replace("µ", "u").replace("μ", "u")
    x = re.sub(r"[^a-z0-9]+", "_", x)
    x = re.sub(r"_+", "_", x)
    return x.strip("_")


def extract_year_from_filename(path):
    """
    Extract year from filename.
    """
    m = re.search(r"(20\d{2})", path.name)
    if m:
        return int(m.group(1))
    return None


def find_excel_files(data_dir):
    """
    Find Excel files in the target folder.
    """
    data_dir = Path(data_dir)

    print("\n============================================================")
    print("CHECKING INPUT FOLDER")
    print("============================================================")
    print("Folder:", data_dir)
    print("Exists:", data_dir.exists())

    if not data_dir.exists():
        raise FileNotFoundError(f"Folder does not exist: {data_dir}")

    print("\nFiles Python can see:")
    for f in sorted(data_dir.iterdir()):
        print(f"  {repr(f.name)} | suffix={repr(f.suffix)} | is_file={f.is_file()}")

    files = []

    for f in data_dir.iterdir():
        if not f.is_file():
            continue

        suffix = f.suffix.lower().strip()
        name_lower = f.name.lower()

        is_excel = suffix in VALID_EXCEL_SUFFIXES
        is_target = "air_quality_ireland" in name_lower

        if is_excel and is_target:
            year = extract_year_from_filename(f)
            if year in YEARS_TO_PROCESS:
                files.append(f)

    files = sorted(files, key=lambda p: extract_year_from_filename(p) or 9999)

    if not files:
        raise FileNotFoundError(
            f"No matching EPA Excel files found in {data_dir}.\n"
            "Expected names like 2020_air_quality_Ireland.xlsm."
        )

    print("\nFiles selected:")
    for f in files:
        print(f"  {f.name}")

    return files


def get_sheet_names(path):
    """
    Return sheet names from workbook.
    """
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


# ============================================================
# 3. HEADER DETECTION AND SHEET PARSING
# ============================================================

def row_nonempty_count(row):
    return sum(clean_text(x) != "" for x in row)


def row_text_fraction(row):
    vals = [clean_text(x) for x in row]
    vals = [v for v in vals if v != ""]

    if len(vals) == 0:
        return 0

    text_like = 0

    for v in vals:
        numeric = pd.to_numeric(pd.Series([v]), errors="coerce").notna().iloc[0]
        dt = pd.to_datetime(pd.Series([v]), errors="coerce", dayfirst=True).notna().iloc[0]

        if not numeric and not dt:
            text_like += 1

    return text_like / len(vals)


def detect_header_rows(raw_df, max_scan_rows=20):
    """
    Detect likely column header row and possible unit row.
    """
    scan = raw_df.head(max_scan_rows).copy()

    candidates = []

    for idx in scan.index:
        row = scan.loc[idx]
        nonempty = row_nonempty_count(row)
        text_frac = row_text_fraction(row)

        if nonempty >= 2 and text_frac >= 0.25:
            candidates.append(idx)

    if candidates:
        header_row = int(candidates[0])
    else:
        header_row = 0

    units_row = None
    header_type = "single_header"

    if header_row + 1 < len(raw_df):
        possible_units = raw_df.iloc[header_row + 1].astype(str).map(clean_text)
        joined = " ".join(possible_units.tolist()).lower()
        joined = joined.replace("µ", "u").replace("μ", "u")

        unit_patterns = [
            r"ug", r"mg", r"g_m", r"g/m", r"m3", r"m\^3",
            r"deg", r"celsius", r"°c", r"%",
            r"hpa", r"mb", r"m/s", r"km/h",
            r"mm", r"w/m", r"ppm", r"ppb"
        ]

        unit_hits = sum(1 for pat in unit_patterns if re.search(pat, joined))
        nonempty_units = sum(
            v != "" and v.lower() not in ["nan", "none"]
            for v in possible_units
        )

        if unit_hits >= 1 and nonempty_units >= 1:
            units_row = header_row + 1
            header_type = "two_level_header_variable_plus_units"

    data_start_row = header_row + 1 if units_row is None else units_row + 1

    return {
        "header_row": header_row,
        "units_row": units_row,
        "data_start_row": data_start_row,
        "header_type": header_type
    }


def make_columns_from_headers(raw_df, header_row, units_row=None):
    """
    Build clean column names.
    If units row exists, preserve units in brackets.
    """
    variables = [clean_text(x) for x in raw_df.iloc[header_row].tolist()]

    if units_row is not None:
        units = [clean_text(x) for x in raw_df.iloc[units_row].tolist()]
    else:
        units = [""] * len(variables)

    cols = []
    seen = {}

    for i, (var, unit) in enumerate(zip(variables, units)):
        if var == "" or var.lower() in ["nan", "none"]:
            var = f"unnamed_col_{i + 1}"

        if unit != "" and unit.lower() not in ["nan", "none", "unit", "units"]:
            col = f"{var} [{unit}]"
        else:
            col = var

        col = clean_text(col)

        if col in seen:
            seen[col] += 1
            col = f"{col}_{seen[col]}"
        else:
            seen[col] = 1

        cols.append(col)

    return cols


def read_sheet_structured(path, sheet_name):
    """
    Read a sheet and parse headers.
    """
    raw = pd.read_excel(
        path,
        sheet_name=sheet_name,
        header=None,
        engine="openpyxl"
    )

    header_info = detect_header_rows(raw)
    cols = make_columns_from_headers(
        raw,
        header_row=header_info["header_row"],
        units_row=header_info["units_row"]
    )

    df = raw.iloc[header_info["data_start_row"]:].copy()
    df.columns = cols

    df = df.dropna(axis=0, how="all")
    df = df.dropna(axis=1, how="all")
    df = df.reset_index(drop=True)

    return df, header_info, raw.shape


# ============================================================
# 4. IDENTIFY PM SHEETS
# ============================================================

def identify_pollutant_from_sheet(sheet_name):
    """
    Identify whether a sheet is PM2.5 or PM10.
    """
    s = normalise_name(sheet_name)

    pm25_patterns = [
        "pm2_5",
        "pm_2_5",
        "pm25",
        "particulate_matter_2_5"
    ]

    pm10_patterns = [
        "pm10",
        "pm_10",
        "particulate_matter_10"
    ]

    if any(p in s for p in pm25_patterns):
        return "PM2.5"

    if any(p in s for p in pm10_patterns):
        return "PM10"

    return None


def identify_timestep_from_sheet(sheet_name):
    """
    Guess timestep from sheet name.
    """
    s = normalise_name(sheet_name)

    if any(x in s for x in ["hourly", "hour", "1hr", "h1"]):
        return "hourly"

    if any(x in s for x in ["daily", "day", "24h", "24_hr", "24hour"]):
        return "daily"

    if any(x in s for x in ["annual", "yearly", "year"]):
        return "annual"

    return "unknown"


def choose_pm_sheets(sheet_names):
    """
    Select PM sheets from workbook.

    Rule:
    - For each pollutant, prefer daily sheets.
    - If daily is absent, allow hourly and aggregate to daily later.
    - Ignore annual summary sheets for this daily dataset.
    """
    candidates = []

    for sheet in sheet_names:
        pollutant = identify_pollutant_from_sheet(sheet)
        timestep = identify_timestep_from_sheet(sheet)

        if pollutant is None:
            continue

        if timestep == "annual":
            continue

        candidates.append({
            "sheet": sheet,
            "pollutant": pollutant,
            "timestep": timestep
        })

    selected = []

    for pollutant in ["PM2.5", "PM10"]:
        pol_candidates = [x for x in candidates if x["pollutant"] == pollutant]

        if not pol_candidates:
            continue

        daily = [x for x in pol_candidates if x["timestep"] == "daily"]
        hourly = [x for x in pol_candidates if x["timestep"] == "hourly"]
        unknown = [x for x in pol_candidates if x["timestep"] == "unknown"]

        if daily:
            selected.extend(daily)
        elif hourly:
            selected.extend(hourly)
        else:
            selected.extend(unknown)

    return selected


# ============================================================
# 5. DATE AND STATION COLUMN DETECTION
# ============================================================

def find_datetime_column(df):
    """
    Find likely date/datetime column.
    """
    # First search by column name
    name_candidates = []

    for col in df.columns:
        c = normalise_name(col)

        if any(x in c for x in ["date", "datetime", "time", "hour", "day"]):
            name_candidates.append(col)

    for col in name_candidates:
        dt = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
        if dt.notna().sum() >= max(5, 0.2 * len(df)):
            return col, dt

    # Fallback: search all columns for datetime-like values
    best_col = None
    best_count = 0
    best_dt = None

    for col in df.columns:
        dt = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
        count = dt.notna().sum()

        if count > best_count:
            best_col = col
            best_count = count
            best_dt = dt

    if best_col is None or best_count < 5:
        return None, None

    return best_col, best_dt


def is_metadata_column(col):
    """
    Exclude obvious non-station columns.
    """
    c = normalise_name(col)

    metadata_terms = [
        "date", "datetime", "time", "hour", "day",
        "year", "month",
        "unnamed",
        "unit", "units",
        "average", "mean", "median",
        "limit", "objective",
        "number_of_valid",
        "valid_data",
        "data_capture",
        "exceedance",
        "exceedances",
        "percentile",
        "max", "min",
        "note", "notes",
        "site_type",
        "station_type",
        "county",
        "zone"
    ]

    return any(term in c for term in metadata_terms)


def get_station_columns(df, datetime_col):
    """
    Identify station concentration columns.

    Assumption:
    - EPA sheets are usually wide:
      date column + one column per station.
    - Station columns should be mostly numeric once coerced.
    """
    station_cols = []

    for col in df.columns:
        if col == datetime_col:
            continue

        if is_metadata_column(col):
            continue

        s = pd.to_numeric(df[col], errors="coerce")
        numeric_count = s.notna().sum()

        # Require at least some numeric values
        if numeric_count >= max(3, 0.05 * len(df)):
            station_cols.append(col)

    return station_cols


# ============================================================
# 6. RESHAPE ONE SHEET TO DAILY LONG FORMAT
# ============================================================

def reshape_pm_sheet_to_daily_long(path, sheet_name, pollutant, timestep_hint, year):
    """
    Parse one PM sheet, reshape to long format, and aggregate to daily if needed.
    """
    df, header_info, raw_shape = read_sheet_structured(path, sheet_name)

    datetime_col, dt = find_datetime_column(df)

    if datetime_col is None:
        raise ValueError("Could not identify date/datetime column.")

    df = df.copy()
    df["_datetime"] = dt

    # Drop rows without valid datetime
    df = df[df["_datetime"].notna()].copy()

    if df.empty:
        raise ValueError("No valid datetime rows after parsing.")

    station_cols = get_station_columns(df, datetime_col)

    if not station_cols:
        raise ValueError("No numeric station concentration columns detected.")

    # Reshape wide to long
    long_df = df.melt(
        id_vars=["_datetime"],
        value_vars=station_cols,
        var_name="station",
        value_name="concentration_ug_m3"
    )

    long_df["concentration_ug_m3"] = pd.to_numeric(
        long_df["concentration_ug_m3"],
        errors="coerce"
    )

    long_df = long_df.dropna(subset=["concentration_ug_m3"]).copy()

    # Remove obvious impossible negative values
    long_df = long_df[long_df["concentration_ug_m3"] >= 0].copy()

    # Daily date
    long_df["date"] = long_df["_datetime"].dt.floor("D")
    long_df["year"] = year
    long_df["pollutant"] = pollutant
    long_df["source_file"] = path.name
    long_df["source_sheet"] = sheet_name
    long_df["source_timestep_hint"] = timestep_hint
    long_df["header_type"] = header_info["header_type"]

    # Aggregate to daily mean.
    # This is harmless if the source is already daily because there should be one value per station-date.
    daily = (
        long_df
        .groupby(
            [
                "date",
                "year",
                "pollutant",
                "station",
                "source_file",
                "source_sheet",
                "source_timestep_hint",
                "header_type"
            ],
            as_index=False
        )
        .agg(
            concentration_ug_m3=("concentration_ug_m3", "mean"),
            n_values_used=("concentration_ug_m3", "count")
        )
    )

    daily = daily[
        [
            "date",
            "year",
            "pollutant",
            "station",
            "concentration_ug_m3",
            "n_values_used",
            "source_file",
            "source_sheet",
            "source_timestep_hint",
            "header_type"
        ]
    ].copy()

    log_row = {
        "file": path.name,
        "year": year,
        "sheet": sheet_name,
        "pollutant": pollutant,
        "timestep_hint": timestep_hint,
        "status": "success",
        "raw_rows": raw_shape[0],
        "raw_cols": raw_shape[1],
        "parsed_rows": df.shape[0],
        "station_cols_detected": len(station_cols),
        "daily_long_rows": daily.shape[0],
        "date_min": daily["date"].min(),
        "date_max": daily["date"].max(),
        "header_type": header_info["header_type"],
        "header_row_excel_like": header_info["header_row"] + 1,
        "units_row_excel_like": None if header_info["units_row"] is None else header_info["units_row"] + 1,
        "data_start_row_excel_like": header_info["data_start_row"] + 1,
        "error": ""
    }

    return daily, log_row


# ============================================================
# 7. MAIN COMBINATION WORKFLOW
# ============================================================

excel_files = find_excel_files(DATA_DIR)

all_daily = []
log_rows = []

print("\n============================================================")
print("PROCESSING WORKBOOKS")
print("============================================================")

for path in excel_files:

    year = extract_year_from_filename(path)

    print("\n------------------------------------------------------------")
    print(f"Workbook: {path.name}")
    print(f"Year: {year}")
    print("------------------------------------------------------------")

    try:
        sheet_names = get_sheet_names(path)
    except Exception as e:
        log_rows.append({
            "file": path.name,
            "year": year,
            "sheet": None,
            "pollutant": None,
            "timestep_hint": None,
            "status": "failed_open_workbook",
            "error": str(e)
        })
        print(f"Failed to open workbook: {e}")
        continue

    if VERBOSE:
        print("Sheets:")
        for s in sheet_names:
            print(f"  - {s}")

    selected_sheets = choose_pm_sheets(sheet_names)

    print("\nSelected PM sheets:")
    if selected_sheets:
        for x in selected_sheets:
            print(f"  - {x['sheet']} | {x['pollutant']} | {x['timestep']}")
    else:
        print("  No PM sheets detected.")

    for x in selected_sheets:
        sheet_name = x["sheet"]
        pollutant = x["pollutant"]
        timestep = x["timestep"]

        print(f"\nReading: {sheet_name} | {pollutant} | {timestep}")

        try:
            daily, log_row = reshape_pm_sheet_to_daily_long(
                path=path,
                sheet_name=sheet_name,
                pollutant=pollutant,
                timestep_hint=timestep,
                year=year
            )

            all_daily.append(daily)
            log_rows.append(log_row)

            print(f"  Success. Daily rows: {daily.shape[0]}")
            print(f"  Date range: {daily['date'].min()} to {daily['date'].max()}")
            print(f"  Stations: {daily['station'].nunique()}")

        except Exception as e:
            print(f"  Failed: {e}")

            log_rows.append({
                "file": path.name,
                "year": year,
                "sheet": sheet_name,
                "pollutant": pollutant,
                "timestep_hint": timestep,
                "status": "failed_parse_sheet",
                "raw_rows": np.nan,
                "raw_cols": np.nan,
                "parsed_rows": np.nan,
                "station_cols_detected": np.nan,
                "daily_long_rows": np.nan,
                "date_min": pd.NaT,
                "date_max": pd.NaT,
                "header_type": None,
                "header_row_excel_like": np.nan,
                "units_row_excel_like": np.nan,
                "data_start_row_excel_like": np.nan,
                "error": str(e)
            })


# ============================================================
# 8. COMBINE AND CLEAN FINAL DATA
# ============================================================

if not all_daily:
    raise RuntimeError("No PM data were successfully parsed. Check the log output above.")

pm_daily = pd.concat(all_daily, ignore_index=True)

# Ensure clean types
pm_daily["date"] = pd.to_datetime(pm_daily["date"])
pm_daily["year"] = pm_daily["year"].astype(int)
pm_daily["pollutant"] = pm_daily["pollutant"].astype(str)
pm_daily["station"] = pm_daily["station"].astype(str)
pm_daily["concentration_ug_m3"] = pd.to_numeric(pm_daily["concentration_ug_m3"], errors="coerce")

# Final cleanup
pm_daily = pm_daily.dropna(subset=["date", "pollutant", "station", "concentration_ug_m3"]).copy()

# If the same date-pollutant-station appears from multiple sheets, average them.
# This protects against duplicate daily/hourly sheets being selected accidentally.
pm_daily_final = (
    pm_daily
    .groupby(["date", "year", "pollutant", "station"], as_index=False)
    .agg(
        concentration_ug_m3=("concentration_ug_m3", "mean"),
        n_values_used=("n_values_used", "sum"),
        source_files=("source_file", lambda x: " | ".join(sorted(set(map(str, x))))),
        source_sheets=("source_sheet", lambda x: " | ".join(sorted(set(map(str, x))))),
        source_timestep_hints=("source_timestep_hint", lambda x: " | ".join(sorted(set(map(str, x)))))
    )
)

pm_daily_final = pm_daily_final.sort_values(
    ["pollutant", "station", "date"]
).reset_index(drop=True)


# ============================================================
# 9. MAKE USEFUL SUMMARY TABLE
# ============================================================

station_summary = (
    pm_daily_final
    .groupby(["year", "pollutant", "station"], as_index=False)
    .agg(
        n_days=("date", "nunique"),
        date_min=("date", "min"),
        date_max=("date", "max"),
        mean_concentration_ug_m3=("concentration_ug_m3", "mean"),
        median_concentration_ug_m3=("concentration_ug_m3", "median"),
        min_concentration_ug_m3=("concentration_ug_m3", "min"),
        max_concentration_ug_m3=("concentration_ug_m3", "max")
    )
    .sort_values(["year", "pollutant", "station"])
)

file_sheet_log = pd.DataFrame(log_rows)


# ============================================================
# 10. SAVE OUTPUTS
# ============================================================

pm_daily_final.to_csv(OUT_LONG_CSV, index=False)
station_summary.to_csv(OUT_SUMMARY_CSV, index=False)
file_sheet_log.to_csv(OUT_LOG_CSV, index=False)

with pd.ExcelWriter(OUT_LONG_XLSX, engine="openpyxl") as writer:
    pm_daily_final.to_excel(writer, sheet_name="pm_daily_long", index=False)
    station_summary.to_excel(writer, sheet_name="station_summary", index=False)
    file_sheet_log.to_excel(writer, sheet_name="file_sheet_log", index=False)


# ============================================================
# 11. FINAL CONSOLE SUMMARY
# ============================================================

print("\n============================================================")
print("COMBINATION COMPLETE")
print("============================================================")

print("\nFinal daily PM dataset:")
print(f"Rows: {pm_daily_final.shape[0]}")
print(f"Stations: {pm_daily_final['station'].nunique()}")
print(f"Pollutants: {sorted(pm_daily_final['pollutant'].unique())}")
print(f"Date range: {pm_daily_final['date'].min()} to {pm_daily_final['date'].max()}")

print("\nRows by year and pollutant:")
print(
    pm_daily_final
    .groupby(["year", "pollutant"])
    .size()
    .reset_index(name="n_rows")
    .to_string(index=False)
)

print("\nNumber of stations by year and pollutant:")
print(
    pm_daily_final
    .groupby(["year", "pollutant"])["station"]
    .nunique()
    .reset_index(name="n_stations")
    .to_string(index=False)
)

print("\nOutputs written:")
print(f"1. {OUT_LONG_CSV}")
print(f"2. {OUT_LONG_XLSX}")
print(f"3. {OUT_SUMMARY_CSV}")
print(f"4. {OUT_LOG_CSV}")

print("\nNext file to inspect:")
print(OUT_LONG_XLSX)
